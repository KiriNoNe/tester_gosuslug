import json
from tkinter import *
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from bs4 import BeautifulSoup
import pandas as pd
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.firefox.service import Service
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import TimeoutException
from tkinter import ttk
from tkinter import messagebox as mb

from threading import Thread
import openpyxl
import os
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment, PatternFill,Border, Side,Font
import calendar
import datetime
import customtkinter

from PIL import Image
class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None

        # Привязываем события
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event):
        # Создаем окно для подсказки, если его ещё нет
        if self.tooltip_window or not self.text:
            return

        x, y, _, _ = self.widget.bbox("insert")  # Координаты элемента
        x += self.widget.winfo_rootx() - 40  # Смещение по оси X
        y += self.widget.winfo_rooty() + 50  # Смещение по оси Y

        # Создаём окно подсказки
        self.tooltip_window = tw = Toplevel(self.widget)
        tw.wm_overrideredirect(True)  # Убираем рамки окна
        tw.geometry(f"+{x}+{y}")

        # Добавляем текст в подсказку
        label = Label(
            tw,
            text=self.text,
            font=("Arial", 10),
            bg="lightyellow",
            relief="solid",
            borderwidth=1,
            padx=5,
            pady=3
        )
        label.pack()

    def hide_tooltip(self, event):
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None
class Window:
    def __init__(self):
        self.root = Tk()  
        self.root.title("MVD")  
        self.root.geometry("720x360")  
        self.root['bg'] = "#EEE2DC"
        Menu(self.root, bg="#505050")
        self.NameSelectedUslugi = []
        self.FullValueSelectedUslugi = []
        
        # Определяем путь к папке с файлами
        self.dir_path = os.path.dirname(os.path.abspath(__file__))
        sub_folder = "files"
        
        # ========== ЗАГРУЗКА УСЛУГ ИЗ JSON ФАЙЛОВ ==========
        
        # Значения по умолчанию для МУ (индекс 0)
        default_uslugiMY = [
            {"text": "Заграничный паспорт", "value": "заграничный_паспорт"},
            {"text": "Внутренний паспорт", "value": "внутренний_паспорт"}
        ]
        self.uslugiMY = self.load_uslugi_from_file("uslugiMY.json", default_uslugiMY)
        
        # Значения по умолчанию для РВП (индекс 1)
        default_uslugiRVP = [
            {"text": "РВП", "value": "рвп"}
        ]
        self.uslugiRVP = self.load_uslugi_from_file("uslugiRVP.json", default_uslugiRVP)
        
        # Значения по умолчанию для ВНЖ (индекс 2)
        default_uslugiVNJ = [
            {"text": "ВНЖ", "value": "внж"}
        ]
        self.uslugiVNJ = self.load_uslugi_from_file("uslugiVNJ.json", default_uslugiVNJ)
        
        # Значения по умолчанию для Патент (индекс 3)
        default_uslugiPATENT = [
            {"text": "Патент", "value": "патент"}
        ]
        self.uslugiPATENT = self.load_uslugi_from_file("uslugiPATENT.json", default_uslugiPATENT)
        
        # Значения по умолчанию для Виза (индекс 4)
        default_uslugiVIZA = [
            {"text": "Виза", "value": "виза"}
        ]
        self.uslugiVIZA = self.load_uslugi_from_file("uslugiVIZA.json", default_uslugiVIZA)
        
        # Значения по умолчанию для РП (индекс 5)
        default_uslugiRP = [
            {"text": "РП", "value": "рп"}
        ]
        self.uslugiRP = self.load_uslugi_from_file("uslugiRP.json", default_uslugiRP)
        
        # Значения по умолчанию для ЗП Нового образца (индекс 6)
        default_uslugiZPNEW = [
            {"text": "Загранпаспорт нового образца", "value": "загранпаспорт_нового_образца"}
        ]
        self.uslugiZPNEW = self.load_uslugi_from_file("uslugiZPNEW.json", default_uslugiZPNEW)
        
        # Значения по умолчанию для ЗП Старого образца (индекс 7)
        default_uslugiZPOLD = [
            {"text": "Загранпаспорт старого образца", "value": "загранпаспорт_старого_образца"}
        ]
        self.uslugiZPOLD = self.load_uslugi_from_file("uslugiZPOLD.json", default_uslugiZPOLD)
        
        # Значения по умолчанию для Приглашение (индекс 8) - СТАТИЧНАЯ
        default_uslugiINVITATION = [
            {"text": "Приглашение", "value": "приглашение"}
        ]
        self.uslugiINVITATION = self.load_uslugi_from_file("uslugiINVITATION.json", default_uslugiINVITATION)
        
        # Значения по умолчанию для АСИ (индекс 9)
        default_uslugiASI = [
            {"text": "АСИ", "value": "аси"}
        ]
        self.uslugiASI = self.load_uslugi_from_file("uslugiASI.json", default_uslugiASI)
        
        # Значения по умолчанию для РУ (индекс 10)
        default_uslugiRY = [
            {"text": "РУ", "value": "ру"}
        ]
        self.uslugiRY = self.load_uslugi_from_file("uslugiRY.json", default_uslugiRY)
        
        # ========== КОНЕЦ ЗАГРУЗКИ ==========
        
        # Загружаем выбранные услуги из JSON (если файл существует)
        json_file_path = os.path.join(self.dir_path, sub_folder, "SelectedUslugi.json")
        if os.path.exists(json_file_path):
            with open(json_file_path, "r", encoding='utf-8') as file:
                selected_data = json.load(file)
                for item in selected_data:
                    self.NameSelectedUslugi.append(item["name"])
                    self.FullValueSelectedUslugi.append(item["value"])
            
            # Проверяем, что загружено 11 элементов
            while len(self.NameSelectedUslugi) < 11:
                self.NameSelectedUslugi.append("Не выбрано")
            while len(self.FullValueSelectedUslugi) < 11:
                self.FullValueSelectedUslugi.append("")
        else:
            # Если файла нет, используем значения по умолчанию
            default_names = [
                "Заграничный паспорт",
                "Внутренний паспорт",
                "ВНЖ",
                "Патент",
                "Виза",
                "РП",
                "Загранпаспорт нового образца",
                "Загранпаспорт старого образца",
                "Приглашение",
                "АСИ",
                "РУ"
            ]
            
            default_values = [
                "заграничный_паспорт",
                "внутренний_паспорт",
                "внж",
                "патент",
                "виза",
                "рп",
                "загранпаспорт_нового_образца",
                "загранпаспорт_старого_образца",
                "приглашение",
                "аси",
                "ру"
            ]
            
            self.NameSelectedUslugi = default_names
            self.FullValueSelectedUslugi = default_values
        
        # Загружаем пути к кнопкам
        file_path_json_path_btn = os.path.join(self.dir_path, sub_folder, "path_to_uslugi_Buttons_name.json")
        if os.path.exists(file_path_json_path_btn):
            with open(file_path_json_path_btn, "r", encoding='utf-8') as file:
                self.data_path_btn = json.load(file)
        else:
            self.data_path_btn = {}


        style = ttk.Style()
        style.configure('Your.TCheckbutton', font=('Trebuchet MS', 12),background="#EEE2DC",foreground='#123C69')
        style.configure('Your.TLabel', font=('Trebuchet MS', 16,'bold'),background="#EEE2DC")    
        style.configure('Your.TRadiobutton',font=('Trebuchet MS', 11),background="#EEE2DC",foreground='#123C69')    
        style.configure('ChoisUs.TLabel', font=('Trebuchet MS', 14,'bold'),background="#EEE2DC")   
        
        self.VarMY = StringVar(value= self.FullValueSelectedUslugi[0])
        self.VarRVP = StringVar(value= self.FullValueSelectedUslugi[1])
        self.VarVNJ = StringVar(value= self.FullValueSelectedUslugi[2])
        self.VarPATENT = StringVar(value= self.FullValueSelectedUslugi[3])
        self.VarVIZA = StringVar(value= self.FullValueSelectedUslugi[4])
        self.VarRP = StringVar(value= self.FullValueSelectedUslugi[5])
        self.VarZPNEW = StringVar(value= self.FullValueSelectedUslugi[6])
        self.VarZPOLD = StringVar(value= self.FullValueSelectedUslugi[7])
        self.VarINVITATION = StringVar(value= self.FullValueSelectedUslugi[8])
        self.VarASI = StringVar(value= self.FullValueSelectedUslugi[9])
        self.VarRY = StringVar(value= self.FullValueSelectedUslugi[10])
        self.listRadioVar = [self.VarMY, self.VarRVP, self.VarVNJ, self.VarPATENT, self.VarVIZA, self.VarRP,
                             self.VarZPNEW, self.VarZPOLD, self.VarINVITATION, self.VarASI, self.VarRY]
        
        self.StringVarMY = StringVar()
        self.StringVarRVP = StringVar()
        self.StringVarVNJ = StringVar()
        self.StringVarPATENT = StringVar()
        self.StringVarVIZA = StringVar()
        self.StringVarRP = StringVar()
        self.StringVarZPNEW = StringVar()
        self.StringVarZPOLD = StringVar()
        self.StringVarINVITATION = StringVar()
        self.StringVarASI = StringVar()
        self.StringVarRY = StringVar()
        
        self.listRadioStringVar =[self.StringVarMY, self.StringVarRVP, self.StringVarVNJ, self.StringVarPATENT
                                  , self.StringVarVIZA, self.StringVarRP,self.StringVarZPNEW, self.StringVarZPOLD
                                  , self.StringVarINVITATION, self.StringVarASI, self.StringVarRY]
        
        
        self.AllUslugi = [self.uslugiMY,self.uslugiRVP,self.uslugiVNJ,self.uslugiPATENT,self.uslugiVIZA,self.uslugiRP,
                          self.uslugiZPNEW,self.uslugiZPOLD,self.uslugiINVITATION,self.uslugiASI,self.uslugiRY]
        
        self.AllProgram = [self.MY,self.RVP,self.VNJ,self.PATENT,self.VIZA,self.replasementPasport,self
                           .ForeignPassportOfANewType,self.ForeignPassportOfAOldType,self.Invitation,self.ASI
                           ,self.RegistrationAtPlaceOfResidence]
        
        self.DataOVDsYVM =[]
        self.DataOVDs = []
        self.nameovdlist= ["MY","RVP","VNJ","PATENT","VIZA","RP","ZPNEW","ZPOLD","INVITATION","ASI","RY"]
        self.nameovdlistru= ["МУ","РВП","ВНЖ","Патент","Виза","РП","ЗП Новый","ЗП Старый","Приглашение","АСИ","РУ"]
        
        self.AllDayinMonth = ['1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
        self.NameOVD = ['Дзержинский','Железнодорожный','Заельцовский','Калининский','Кировский','Ленинский','Октябрьский','Первомайский','Советский'
                        ,'Центральный','Бердск','Искитимский','Отдел по вопросам миграции Куйбышевский','Мошково'
                        ,'Отдел  по вопросам миграции Новосибирский','Кольцово','Обь','Краснообск','Маслянино','Колывань'
                        ,'Отдел по вопросам миграции Барабинск','Баган','Убинское','Отдел по вопросам миграции Карасук'
                        , 'Чулым','Отдел по вопросам миграции Краснозерский','Доволенское','Болотное'
                        ,'Миграционный пункт межмуниципального отдела МВД России Каргатский','Коченево','Кочки'
                        ,'Отдел по вопросам миграции Татарский','Чаны','Купино','межмуниципального отдела МВД России "Венгеровский"'
                        ,'Здвинское','Кыштовское','Ордынское','Северное','Сузун','Тогучин','Усть-Тарка','Черепаново','Чистоозерное']
        
        self.NameOVDReplasement = ['Дзержинский','Железнодорожный','Заельцовский','Калининский','Кировский','Ленинский','Октябрьский','Первомайский','Советский','Центральный','Бердск','Искитимский','Отдел по вопросам миграции Куйбышевский','Мошково','Отдел  по вопросам миграции Новосибирский','Кольцово','Обь','Краснообск','Маслянино','Колывань','Отдел по вопросам миграции Барабинск','Баган','Убинское','Отдел по вопросам миграции Карасук','Чулым','Отдел по вопросам миграции Краснозерский','Доволенское','Болотное','Миграционный пункт межмуниципального отдела МВД России Каргатский','Коченево','Кочки','Отдел по вопросам миграции Татарский','Чаны','Купино','межмуниципального отдела МВД России "Венгеровский"','Здвинское','Кыштовское','Ордынское','Северное','Сузун','Тогучин','Усть-Тарка','Черепаново','Чистоозерное']
        self.NameOVDPassNewType = ['Железнодорожный','Ленинский','Октябрьский','Первомайский','Советский','Искитимский','Отдел по вопросам миграции Куйбышевский','Мошково','Обь','Отдел по вопросам миграции Карасук','Отдел по вопросам миграции Татарский']
        self.NameOVDInvitation = ['Дзержинский','Железнодорожный','Заельцовский','Калининский','Кировский','Ленинский','Октябрьский','Первомайский','Советский','Центральный','Бердск','Искитимский','Отдел по вопросам миграции Куйбышевский','Мошково','Отдел  по вопросам миграции Новосибирский','Кольцово','Обь','Маслянино','Колывань','Отдел по вопросам миграции Барабинск','Баган','Убинское','Отдел по вопросам миграции Карасук','Чулым','Отдел по вопросам миграции Краснозерский','Болотное','Доволенское','Миграционный пункт межмуниципального отдела МВД России Каргатский','Коченево','Кочки','Отдел по вопросам миграции Татарский','Чаны','Купино','межмуниципального отдела МВД России "Венгеровский"','Здвинское','Кыштовское','Ордынское','Северное','Сузун','Тогучин','Усть-Тарка','Черепаново','Чистоозерное']
        self.NameOVDPassOldType = ['Дзержинский','Железнодорожный','Заельцовский','Калининский','Кировский','Ленинский','Октябрьский','Первомайский','Советский','Центральный','Бердск','Искитимский','Отдел по вопросам миграции Куйбышевский','Мошково','Отдел  по вопросам миграции Новосибирский','Кольцово','Обь','Краснообск','Маслянино','Колывань','Отдел по вопросам миграции Барабинск','Баган','Убинское','Отдел по вопросам миграции Карасук','Чулым','Отдел по вопросам миграции Краснозерский','Доволенское','Болотное','Миграционный пункт межмуниципального отдела МВД России Каргатский','Коченево','Кочки','Отдел по вопросам миграции Татарский','Чаны','Купино','межмуниципального отдела МВД России "Венгеровский"','Здвинское','Кыштовское','Ордынское','Северное','Сузун','Тогучин','Усть-Тарка','Черепаново','Чистоозерное']
        self.NameOVDASI = ['Дзержинский','Железнодорожный','Заельцовский','Калининский','Кировский','Ленинский','Октябрьский','Первомайский','Советский','Центральный','Бердск','Искитимский','Отдел по вопросам миграции Куйбышевский','Мошково','Отдел  по вопросам миграции Новосибирский','Кольцово','Обь','Краснообск','Маслянино','Колывань','Отдел по вопросам миграции Барабинск','Баган','Убинское','Отдел по вопросам миграции Карасук','Чулым','Отдел по вопросам миграции Краснозерский','Доволенское','Болотное','Миграционный пункт межмуниципального отдела МВД России Каргатский','Коченево','Кочки','Отдел по вопросам миграции Татарский','Чаны','Купино','межмуниципального отдела МВД России "Венгеровский"','Здвинское','Кыштовское','Ордынское','Северное','Сузун','Тогучин','Усть-Тарка','Черепаново','Чистоозерное']
        self.NameOVDRegistrationAtPlaceOfResidence = ['Дзержинский','Железнодорожный','Заельцовский','Калининский','Кировский','Ленинский','Октябрьский','Первомайский','Советский','Центральный','Бердск','Искитимский','Отдел по вопросам миграции Куйбышевский','Мошково','Отдел  по вопросам миграции Новосибирский','Кольцово','Обь','Краснообск','Маслянино','Колывань','Отдел по вопросам миграции Барабинск','Баган','Убинское','Отдел по вопросам миграции Карасук','Чулым','Отдел по вопросам миграции Краснозерский','Доволенское','Болотное','Миграционный пункт межмуниципального отдела МВД России Каргатский','Коченево','Кочки','Отдел по вопросам миграции Татарский','Чаны','Купино','межмуниципального отдела МВД России "Венгеровский"','Здвинское','Кыштовское','Ордынское','Северное','Сузун','Тогучин','Усть-Тарка','Черепаново','Чистоозерное']
        #self.NameOVDGPPS = ['Дзержинский','Железнодорожный','Заельцовский','Калининский','Кировский','Ленинский','Октябрьский','Первомайский','Советский','Центральный','Бердск','Искитимский','Отдел по вопросам миграции Куйбышевский','Мошково','Отдел  по вопросам миграции Новосибирский','Кольцово','Обь','Маслянино','Колывань','Отдел по вопросам миграции Барабинск','Баган','Убинское','Отдел по вопросам миграции Карасук','Чулым','Отдел по вопросам миграции Краснозерский','Болотное','Доволенское','Миграционный пункт межмуниципального отдела МВД России Каргатский','Коченево','Кочки','Отдел по вопросам миграции Татарский','Чаны','Купино','межмуниципального отдела МВД России "Венгеровский"','Здвинское','Кыштовское','Ордынское','Северное','Сузун','Тогучин','Усть-Тарка','Черепаново','Чистоозерное']
        self.NameOVDMY = ['Дзержинский','Железнодорожный','Заельцовский','Калининский','Кировский','Ленинский','Октябрьский','Первомайский','Советский','Центральный','Бердск','Искитимский','Отдел по вопросам миграции Куйбышевский','Мошково','Отдел  по вопросам миграции Новосибирский','Кольцово','Обь','Краснообск','Маслянино','Колывань','Отдел по вопросам миграции Барабинск','Баган','Убинское','Отдел по вопросам миграции Карасук','Чулым','Отдел по вопросам миграции Краснозерский','Доволенское','Болотное','Миграционный пункт межмуниципального отдела МВД России Каргатский','Коченево','Кочки','Отдел по вопросам миграции Татарский','Чаны','Купино','межмуниципального отдела МВД России "Венгеровский"','Здвинское','Кыштовское','Ордынское','Северное','Сузун','Тогучин','Усть-Тарка','Черепаново','Чистоозерное']
        self.NameOVDRVP = ['Бердск','Искитимский','Отдел по вопросам миграции Куйбышевский','Мошково','Обь','Маслянино','Колывань','Отдел по вопросам миграции Барабинск','Баган','Убинское','Отдел по вопросам миграции Карасук','Чулым','Отдел по вопросам миграции Краснозерский','Доволенское','Болотное','Миграционный пункт межмуниципального отдела МВД России Каргатский','Коченево','Кочки','Отдел по вопросам миграции Татарский','Чаны','Купино','межмуниципального отдела МВД России "Венгеровский"','Здвинское','Кыштовское','Ордынское','Северное','Сузун','Тогучин','Усть-Тарка','Черепаново','Чистоозерное']
        self.NameOVDVNJ = ['Бердск','Искитимский','Отдел по вопросам миграции Куйбышевский','Мошково','Обь','Маслянино','Колывань','Отдел по вопросам миграции Барабинск','Баган','Убинское','Отдел по вопросам миграции Карасук','Чулым','Отдел по вопросам миграции Краснозерский','Доволенское','Болотное','Миграционный пункт межмуниципального отдела МВД России Каргатский','Коченево','Кочки','Отдел по вопросам миграции Татарский','Чаны','Купино','межмуниципального отдела МВД России "Венгеровский"','Здвинское','Кыштовское','Ордынское','Северное','Сузун','Тогучин','Усть-Тарка','Черепаново','Чистоозерное']
        self.NameOVDPATENT = ['Маслянино','Отдел по вопросам миграции Барабинск','Баган','Убинское','Отдел по вопросам миграции Карасук','Чулым','Отдел по вопросам миграции Краснозерский','Доволенское','Болотное','Миграционный пункт межмуниципального отдела МВД России Каргатский','Кочки','Отдел по вопросам миграции Татарский','Чаны','Купино','межмуниципального отдела МВД России "Венгеровский"','Здвинское','Кыштовское','Ордынское','Северное','Сузун','Тогучин','Усть-Тарка','Черепаново','Чистоозерное']
        self.NameOVDVIZA = []
        
        self.NameOVDMVD = ['Управление по вопросам миграции','Есенина','Дуси Ковальчук']
        self.NameOVDReplasementMVD = []
        self.NameOVDPassNewTypeMVD = ['Управление по вопросам миграции','Дуси Ковальчук']
        self.NameOVDInvitationMVD = ['Управление по вопросам миграции']
        self.NameOVDPassOldTypeMVD = ['Управление по вопросам миграции']
        self.NameOVDASIMVD = []
        self.NameOVDRegistrationAtPlaceOfResidenceMVD = []
        #self.NameOVDGPPSMVD = ['Управление по вопросам миграции']
        self.NameOVDMYMVD = []
        self.NameOVDRVPMVD = ['Есенина']
        self.NameOVDVNJMVD = ['Есенина']
        self.NameOVDPATENTMVD = ['Есенина']
        self.NameOVDVIZAMVD = ['Управление по вопросам миграции']
        
        self.month_translation = {
                        'January': 'января',
                        'February': 'февраля',
                        'March': 'марта',
                        'April': 'апреля',
                        'May': 'мая',
                        'June': 'июня',
                        'July': 'июля',
                        'August': 'августа',
                        'September': 'сентября',
                        'October': 'октября',
                        'November': 'ноября',
                        'December': 'декабря'
                    }
        
        self.holidays = [
            datetime.date(datetime.datetime.now().year, 1, 1),
            datetime.date(datetime.datetime.now().year, 1, 2),
            datetime.date(datetime.datetime.now().year, 1, 3),
            datetime.date(datetime.datetime.now().year, 1, 4),
            datetime.date(datetime.datetime.now().year, 1, 5),
            datetime.date(datetime.datetime.now().year, 1, 6),
            datetime.date(datetime.datetime.now().year, 1, 7),
            datetime.date(datetime.datetime.now().year, 1, 8),
            datetime.date(datetime.datetime.now().year, 2, 23),
            datetime.date(datetime.datetime.now().year, 3, 8),
            datetime.date(datetime.datetime.now().year, 5, 1),
            datetime.date(datetime.datetime.now().year, 5, 9),
            datetime.date(datetime.datetime.now().year, 6, 12),
            datetime.date(datetime.datetime.now().year, 11, 4),
                    ]
        
        sub_folder = "files"
        self.sub_folder = "files"
        #for i,ChooseUs in enumerate(self.nameovdlist):
        #    file_path = os.path.join(self.dir_path, sub_folder, f"ChooseUslugi{ChooseUs}.txt")
        #    with open(file_path, "r+", encoding='utf-8') as file:
        #        lines = file.read().splitlines()
        #    for line in lines:
        #        value, text = line.split(':')
        #        ButtonDict = {}
        #        ButtonDict["value"] = value
        #        ButtonDict["text"] = text
        #        eval(f"self.uslugi{ChooseUs}.append(ButtonDict)")
        
        self.indexDF = ['Дзержинский','Железнодорожный','Заельцовский','Калининский','Кировский','Ленинский','Октябрьский'
                   ,'Первомайский','Советский','Центральный','Бердск','Искитимский','Куйбышевский','Мошково','Новосибирский'
                   ,'Кольцово','Обь','Краснообск','Маслянино','Колывань','Барабинск','Баган','Убинское','Карасук','Чулым'
                    ,'Краснозерское','Доволенское','Болотное',"Каргат",'Коченево','Кочки','Татарск','Чаны','Купино','Венгерово'
                        ,'Здвинск','Кыштовка','Ордынское','Северное','Сузун','Тогучин','Усть-Тарка','Черепаново','Чистоозерное']
        self.indexDFMVD = ['УВМ','УВМ ОВТМ','УВМ Дуси Ковальчук']
    
    def load_uslugi_from_file(self, filename, default_uslugi):
        """Загружает список услуг из JSON файла"""
        sub_folder = "files"
        file_path = os.path.join(self.dir_path, sub_folder, filename)
        
        if os.path.exists(file_path):
            with open(file_path, "r", encoding='utf-8') as file:
                return json.load(file)
        else:
            # Сохраняем значения по умолчанию
            with open(file_path, "w", encoding='utf-8') as file:
                json.dump(default_uslugi, file, ensure_ascii=False, indent=4)
            return default_uslugi

    def StartBrowser(self):
        self.firefox_path = os.path.join(self.dir_path, "firefox", "firefox")
        file_path = os.path.join(self.dir_path,"geckodriver")
        service = Service(executable_path=file_path)
        options = webdriver.FirefoxOptions()
        options.binary_location = self.firefox_path
        self.browser = webdriver.Firefox(service=service, options=options)
        self.browser.get('https://www.gosuslugi.ru/600300/1/form')
        self.wait = WebDriverWait(self.browser,20,poll_frequency=2)
    
    
    def create_data_frames(self):
        self.dfMVD = pd.DataFrame({f"МУ ({self.NameSelectedUslugi[0]})":[''],f"РВП ({self.NameSelectedUslugi[1]})":['']
                              ,f"ВНЖ ({self.NameSelectedUslugi[2]})":[''],f"ПАТЕНТ ({self.NameSelectedUslugi[3]})":['']
                              ,f"ВИЗА ({self.NameSelectedUslugi[4]})":[''],f'РП ({self.NameSelectedUslugi[5]})':['']
                              ,f"ЗП НОВОГО ОБРАЗЦА ({self.NameSelectedUslugi[6]})":['']
                              ,f"ЗП СТАРОГО ОБРАЗЦА ({self.NameSelectedUslugi[7]})":['']
                              ,f"ПРИГЛАШЕНИЕ ({self.NameSelectedUslugi[8]})":[''],f"АСИ ({self.NameSelectedUslugi[9]})":['']
                              ,f"РУ гр. РФ по месту жит-ва ({self.NameSelectedUslugi[10]})":['']},self.indexDFMVD)
        self.df = pd.DataFrame({f"МУ ({self.NameSelectedUslugi[0]})":[''],f"РВП ({self.NameSelectedUslugi[1]})":['']
                              ,f"ВНЖ ({self.NameSelectedUslugi[2]})":[''],f"ПАТЕНТ ({self.NameSelectedUslugi[3]})":['']
                              , f"ВИЗА ({self.NameSelectedUslugi[4]})":[''],f'РП ({self.NameSelectedUslugi[5]})':['']
                              ,f"ЗП НОВОГО ОБРАЗЦА ({self.NameSelectedUslugi[6]})":['']
                              ,f"ЗП СТАРОГО ОБРАЗЦА ({self.NameSelectedUslugi[7]})":['']
                              ,f"ПРИГЛАШЕНИЕ ({self.NameSelectedUslugi[8]})":[''],f"АСИ ({self.NameSelectedUslugi[9]})":['']
                              ,f"РУ гр. РФ по месту жит-ва ({self.NameSelectedUslugi[10]})":['']},self.indexDF)
        
        
    def run(self):
        self.create_data_frames()
        self.draw_window()
        self.root.mainloop()
    def start_thread(self, func):
        thread = Thread(target=func)
        thread.start()
        thread.join()
    def draw_window(self):

        for c in range(7): self.root.columnconfigure(index=c, weight=1)
        for r in range(5): self.root.rowconfigure(index=r, weight=1)

        btnStartBrowser = customtkinter.CTkButton(self.root, text="Запустить браузер",font=('Trebuchet MS', 16,'bold'),
                                                  fg_color='#EDC7B7',hover_color='#A58B80',text_color='#AC3B61', 
                                                  command=lambda: Thread(target=self.StartBrowser).start())
        btnStartBrowser.grid(row=1, column=1,sticky='WENS')
        ClockChooseUslugi = self.ClockChooseUslugi
        btnChooseUslugi = customtkinter.CTkButton(self.root, text="Выбрать цели",font=('Trebuchet MS', 16,'bold'),
                                                  fg_color='#EDC7B7',hover_color='#A58B80',text_color='#AC3B61', 
                                                  command=lambda: Thread(target=self.ClockChooseUslugi).start())
        btnChooseUslugi.grid(row=1, column=3,sticky='WENS')

        btnStartProgram = customtkinter.CTkButton(self.root, text="Запуск программы гражданина",
                                                  font=('Trebuchet MS', 16,'bold'),fg_color='#EDC7B7',hover_color='#A58B80',
                                                  text_color='#AC3B61',command=lambda: Thread(target=self.StartProgramRP).start())
        btnStartProgram.grid(row=1, column=5,sticky='WENS')
        
        btnStartProgram = customtkinter.CTkButton(self.root, text="Запуск программы мигранта",
                                                  font=('Trebuchet MS', 16,'bold'),fg_color='#EDC7B7',hover_color='#A58B80',
                                                  text_color='#AC3B61',command=lambda: Thread(target=self.StartProgramImigrant).start())
        btnStartProgram.grid(row=3, column=5,sticky='WENS')
        
        btnViewTable = customtkinter.CTkButton(self.root, text="Показать таблицу",font=('Trebuchet MS', 16,'bold'),
                                               fg_color='#EDC7B7',hover_color='#A58B80',text_color='#AC3B61',
                                               command=lambda: Thread(target=self.ViewTable).start())
        btnViewTable.grid(row=3, column=1,sticky='WENS')
        
        btnStartBrowser = customtkinter.CTkButton(self.root, text="Выбрать запуск услуг",font=('Trebuchet MS', 16,'bold'),
                                                  fg_color='#EDC7B7',hover_color='#A58B80',text_color='#AC3B61', 
                                                  command=lambda: Thread(target=self.StartOfIndividualServices).start())
        btnStartBrowser.grid(row=3, column=3,sticky='WENS')

        sub_folder = "icon"
        image_path = os.path.join(self.dir_path, sub_folder, "icon-setting.png")
        button_image = customtkinter.CTkImage(
            Image.open(image_path),  # Изображение 
            size=(15, 15)  # Размер изображения (можно изменить)
        )
        btnSetting = customtkinter.CTkButton(self.root, text="",image=button_image,fg_color='#EDC7B7',   
                                            hover_color='#A58B80',  width=30,  height=30,
                                            command=lambda: Thread(target=self.StartSetting).start())
        btnSetting.grid(row=4, column=3,sticky='')
    
    def save_path_name_btn(self, name_usluga, entry_name_btn):
        """Сохраняет данные в JSON файл."""
        self.data_path_btn[name_usluga] = entry_name_btn
        file_path_json_path_btn = os.path.join(self.dir_path, "files", "path_to_uslugi_Buttons_name.json")
        with open(file_path_json_path_btn, "w", encoding="utf-8") as file:
            json.dump(self.data_path_btn, file, indent=4, ensure_ascii=False)

    def StartSetting(self):
        """Создаёт окно настроек."""
        self.window_setting = Toplevel(self.root)
        self.window_setting.title("Настройки")
        self.window_setting.geometry("600x600")  # Увеличил высоту для новых кнопок
        self.window_setting["bg"] = "#EEE2DC"

        # Настройка колонок и строк
        for c in range(5):
            self.window_setting.columnconfigure(index=c, weight=1)
        for r in range(10):  # Увеличил количество строк
            self.window_setting.rowconfigure(index=r, weight=1)

        # ----- Существующие настройки (Путь к услугам) -----
        label = ttk.Label(self.window_setting, text="Путь к услугам", foreground="#AC3B61", style="Your.TLabel", anchor="center")
        label.grid(row=1, column=1, columnspan=1)
        btn = customtkinter.CTkButton(
                self.window_setting,
                text="Перейти",
                font=("Trebuchet MS", 14, "bold"),
                fg_color="#EDC7B7",
                hover_color="#A58B80",
                text_color="#AC3B61",
                command=lambda: Thread(
                    target=self.start_setting_path_uslugi
                ).start()
            )
        btn.grid(row=2, column=1)

        # ----- Существующие настройки (Путь к подуслугам) -----
        #label = ttk.Label(self.window_setting, text="Путь к подуслугам", foreground="#AC3B61", style="Your.TLabel", anchor="center")
        #label.grid(row=1, column=3, columnspan=1)
        #btn = customtkinter.CTkButton(
        #        self.window_setting,
        #        text="Перейти",
        #        font=("Trebuchet MS", 14, "bold"),
        #        fg_color="#EDC7B7",
        #        hover_color="#A58B80",
        #        text_color="#AC3B61",
        #        command=lambda: Thread(
        #            target=self.start_setting_path_poduslugi
        #        ).start()
        #    )
        #btn.grid(row=2, column=3)

        # ========== НОВАЯ СЕКЦИЯ: Редактирование списка услуг ==========
        # Разделитель (визуальный)
        separator = ttk.Separator(self.window_setting, orient='horizontal')
        separator.grid(row=4, column=0, columnspan=5, sticky='ew', pady=10)
        
        label = ttk.Label(
            self.window_setting, 
            text="Редактирование списка услуг (радио-кнопки)", 
            foreground="#AC3B61", 
            style="Your.TLabel", 
            anchor="center"
        )
        label.grid(row=5, column=1, columnspan=3, pady=10)
        
        # Кнопка для редактирования списка услуг
        btn_edit_services = customtkinter.CTkButton(
            self.window_setting,
            text="Редактировать список услуг",
            font=("Trebuchet MS", 14, "bold"),
            fg_color="#EDC7B7",
            hover_color="#A58B80",
            text_color="#AC3B61",
            command=lambda: Thread(target=self.start_setting_edit_services_list).start()
        )
        btn_edit_services.grid(row=6, column=1, columnspan=1)
        
        # Кнопка для сброса к значениям по умолчанию
        btn_reset_services = customtkinter.CTkButton(
            self.window_setting,
            text="Сбросить к значениям по умолчанию",
            font=("Trebuchet MS", 12, "bold"),
            fg_color="#EDC7B7",
            hover_color="#A58B80",
            text_color="#AC3B61",
            command=lambda: Thread(target=self.reset_services_to_default).start()
        )
        btn_reset_services.grid(row=6, column=3, columnspan=1)
        
        # Пояснение
        tooltip_label = ttk.Label(
            self.window_setting,
            text="Изменение названий и значений услуг,\nкоторые отображаются в окне 'Выбрать цели'",
            foreground="#666666",
            font=("Trebuchet MS", 9),
            background="#EEE2DC",
            justify="center"
        )
        tooltip_label.grid(row=7, column=1, columnspan=3, pady=10)
    
    def start_setting_edit_services_list(self):
        """Окно для редактирования всех услуг (радио-кнопок)"""
        window_edit = Toplevel(self.window_setting)
        window_edit.title("Редактирование списка услуг")
        window_edit.geometry("1100x800")
        window_edit["bg"] = "#EEE2DC"
        
        # Создаем canvas с прокруткой для большого списка
        canvas = Canvas(window_edit, bg="#EEE2DC", highlightthickness=0)
        scrollbar = ttk.Scrollbar(window_edit, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Заголовки
        headers = ["Категория", "№", "Текст (отображается)", "Значение (value)", "Действия"]
        for col, header in enumerate(headers):
            ttk.Label(
                scrollable_frame,
                text=header,
                foreground="#AC3B61",
                font=("Trebuchet MS", 12, "bold"),
                background="#EEE2DC",
                padding=5
            ).grid(row=0, column=col, padx=5, pady=5, sticky="w")
        
        # Список всех категорий услуг
        all_categories = [
            ("МУ", "MY", self.uslugiMY),
            ("РВП", "RVP", self.uslugiRVP),
            ("ВНЖ", "VNJ", self.uslugiVNJ),
            ("Патент", "PATENT", self.uslugiPATENT),
            ("Виза", "VIZA", self.uslugiVIZA),
            ("РП", "RP", self.uslugiRP),
            ("ЗП Нового образца", "ZPNEW", self.uslugiZPNEW),
            ("ЗП Старого образца", "ZPOLD", self.uslugiZPOLD),
            ("Приглашение", "INVITATION", self.uslugiINVITATION),
            ("АСИ", "ASI", self.uslugiASI),
            ("РУ", "RY", self.uslugiRY),
        ]
        
        self.edit_vars = []  # Сохраняем переменные для сохранения
        
        row = 1
        for category_name, category_key, uslugi_list in all_categories:
            if not uslugi_list:
                continue
                
            for idx, usluga in enumerate(uslugi_list):
                # Категория (объединяем ячейки)
                if idx == 0:
                    ttk.Label(
                        scrollable_frame,
                        text=category_name,
                        foreground="#d77991",
                        font=("Trebuchet MS", 10, "bold"),
                        background="#EEE2DC",
                        padding=5
                    ).grid(row=row, column=0, padx=5, pady=3, sticky="nw")
                else:
                    ttk.Label(
                        scrollable_frame,
                        text="",
                        background="#EEE2DC"
                    ).grid(row=row, column=0, padx=5, pady=3)
                
                # Номер
                ttk.Label(
                    scrollable_frame,
                    text=str(idx + 1),
                    background="#EEE2DC",
                    padding=5
                ).grid(row=row, column=1, padx=5, pady=3)
                
                # Поле для текста (отображаемого)
                text_var = StringVar(value=usluga.get("text", ""))
                text_entry = customtkinter.CTkEntry(
                    scrollable_frame,
                    textvariable=text_var,
                    width=350
                )
                text_entry.grid(row=row, column=2, padx=5, pady=3)
                
                # Поле для значения (value)
                value_var = StringVar(value=usluga.get("value", ""))
                value_entry = customtkinter.CTkEntry(
                    scrollable_frame,
                    textvariable=value_var,
                    width=250
                )
                value_entry.grid(row=row, column=3, padx=5, pady=3)
                
                # Сохраняем информацию для сохранения
                self.edit_vars.append({
                    "category_key": category_key,
                    "category_name": category_name,
                    "uslugi_list": uslugi_list,
                    "index": idx,
                    "text_var": text_var,
                    "value_var": value_var
                })
                
                # Кнопка "Сохранить" для отдельной строки
                save_btn = customtkinter.CTkButton(
                    scrollable_frame,
                    text="Сохранить",
                    font=("Trebuchet MS", 10, "bold"),
                    fg_color="#EDC7B7",
                    hover_color="#A58B80",
                    text_color="#AC3B61",
                    width=80,
                    command=lambda item=self.edit_vars[-1]: self.save_single_service(item)
                )
                save_btn.grid(row=row, column=4, padx=5, pady=3)
                
                row += 1
        
        # Кнопка "Сохранить всё"
        save_all_btn = customtkinter.CTkButton(
            scrollable_frame,
            text="Сохранить все изменения",
            font=("Trebuchet MS", 14, "bold"),
            fg_color="#EDC7B7",
            hover_color="#A58B80",
            text_color="#AC3B61",
            command=self.save_all_services_changes
        )
        save_all_btn.grid(row=row, column=1, columnspan=2, pady=20)
        
        # Кнопка закрытия
        close_btn = customtkinter.CTkButton(
            scrollable_frame,
            text="Закрыть",
            font=("Trebuchet MS", 12, "bold"),
            fg_color="#EDC7B7",
            hover_color="#A58B80",
            text_color="#AC3B61",
            command=window_edit.destroy
        )
        close_btn.grid(row=row, column=3, columnspan=1, pady=20)
        
        # Упаковка canvas и scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    
    def save_single_service(self, item):
        """Сохраняет одну услугу"""
        uslugi_list = item["uslugi_list"]
        idx = item["index"]
        new_text = item["text_var"].get()
        new_value = item["value_var"].get()
        
        uslugi_list[idx]["text"] = new_text
        uslugi_list[idx]["value"] = new_value
        
        # Сохраняем в файл
        filename = f"uslugi{item['category_key']}.json"
        self.save_uslugi_to_file(uslugi_list, filename)
        
        mb.showinfo("Успех", f"Сохранено: {item['category_name']} - {new_text}")

    def save_all_services_changes(self):
        """Сохраняет все изменения"""
        for item in self.edit_vars:
            uslugi_list = item["uslugi_list"]
            idx = item["index"]
            new_text = item["text_var"].get()
            new_value = item["value_var"].get()
            
            uslugi_list[idx]["text"] = new_text
            uslugi_list[idx]["value"] = new_value
        
        # Сохраняем каждую категорию в свой файл
        categories = {
            "MY": self.uslugiMY,
            "RVP": self.uslugiRVP,
            "VNJ": self.uslugiVNJ,
            "PATENT": self.uslugiPATENT,
            "VIZA": self.uslugiVIZA,
            "RP": self.uslugiRP,
            "ZPNEW": self.uslugiZPNEW,
            "ZPOLD": self.uslugiZPOLD,
            "INVITATION": self.uslugiINVITATION,
            "ASI": self.uslugiASI,
            "RY": self.uslugiRY,
        }
        
        for key, uslugi_list in categories.items():
            if uslugi_list:
                filename = f"uslugi{key}.json"
                self.save_uslugi_to_file(uslugi_list, filename)
        
        mb.showinfo("Успех", "Все изменения сохранены!")

    def save_uslugi_to_file(self, uslugi_list, filename):
        """Сохраняет список услуг в JSON файл"""
        sub_folder = "files"
        file_path = os.path.join(self.dir_path, sub_folder, filename)
        with open(file_path, "w", encoding='utf-8') as file:
            json.dump(uslugi_list, file, ensure_ascii=False, indent=4)

    def reset_services_to_default(self):
        """Сбрасывает все услуги к значениям по умолчанию"""
        if mb.askyesno("Подтверждение", "Вы уверены? Все изменения будут потеряны!"):
            # Здесь нужно определить значения по умолчанию для каждой категории
            # и перезаписать файлы
            mb.showinfo("Успех", "Список услуг сброшен к значениям по умолчанию")
            # После сброса нужно перезагрузить данные
            self.load_all_services()

    
    def start_setting_path_uslugi(self):
        """Создаёт окно настроек."""
        window_setting_uslugi = Toplevel(self.window_setting)
        window_setting_uslugi.title("Настройки")
        window_setting_uslugi.geometry("600x800")
        window_setting_uslugi["bg"] = "#EEE2DC"
        for c in range(4):
            window_setting_uslugi.columnconfigure(index=c, weight=1)
        for r in range(14):
            window_setting_uslugi.rowconfigure(index=r, weight=1)

        # Заголовок
        label = ttk.Label(window_setting_uslugi, text="Путь к услугам", foreground="#AC3B61", style="Your.TLabel", anchor="center")
        label.grid(row=1, column=2, columnspan=1)

        ToolTip(label, """Если услуга перестала работать, это может означать, что текст на кнопке,
                      ведущей к услуге (например, на кнопку перехода к паспорту РФ), был изменен.
                      В таком случае необходимо обновить этот текст в приложении для 
                      восстановления корректной работы функции.""")

        # Список услуг и привязка StringVar
        services = {
            "МУ": "my", "РВП": "rvp", "ВНЖ": "vnj", "Патент": "patent", "Виза": "viza",
            "РП": "rp", "ЗП": "zp", "Приглашение": "invitation", "АСИ": "asi", "РУ": "ry"
        }
        entry_vars = {}

        for i, (label_text, service_key) in enumerate(services.items()):
            # Метки
            label = ttk.Label(
                window_setting_uslugi, text=label_text, foreground="#d77991", font=("Trebuchet MS", 12, "bold"),
                background="#EEE2DC", anchor="center"
            )
            label.grid(row=2 + i, column=1)

            # Поля ввода
            entry_var = customtkinter.StringVar()
            entry_var.set(self.data_path_btn.get(service_key, ""))  # Получаем начальное значение
            entry_vars[service_key] = entry_var  # Сохраняем StringVar для каждого сервиса

            entry = customtkinter.CTkEntry(window_setting_uslugi, textvariable=entry_var, width=300)
            entry.grid(row=2 + i, column=2)

            # Кнопки "Сохранить"
            btn = customtkinter.CTkButton(
                window_setting_uslugi,
                text="Сохранить",
                font=("Trebuchet MS", 14, "bold"),
                fg_color="#EDC7B7",
                hover_color="#A58B80",
                text_color="#AC3B61",
                command=lambda key=service_key, var=entry_var: Thread(
                    target=self.save_path_name_btn, args=(key, var.get())
                ).start()
            )
            btn.grid(row=2 + i, column=4)

    def start_setting_path_poduslugi(self):
        window_setting_poduslugi = Toplevel(self.window_setting)
        window_setting_poduslugi.title("Настройки")
        window_setting_poduslugi.geometry("600x800")
        window_setting_poduslugi["bg"] = "#EEE2DC"
        for c in range(4):
            window_setting_poduslugi.columnconfigure(index=c, weight=1)
        for r in range(14):
            window_setting_poduslugi.rowconfigure(index=r, weight=1)

        label = ttk.Label(window_setting_poduslugi, text="Выберите услугу", foreground="#AC3B61", style="Your.TLabel", anchor="center")
        label.grid(row=1, column=1, columnspan=1)
        services = {
            "МУ": "MY", "РВП": "RVP", "ВНЖ": "VNJ", "Патент": "PATENT", "Виза": "VIZA",
            "РП": "RP", "ЗП НОВЫЙ": "ZPNEW","ЗП СТАРЫЙ": "ZPOLD", "Приглашение": "INVITATION", "АСИ": "ASI", "РУ": "RY"
        }
        for i, (label_text, service_key) in enumerate(services.items()):
            # Метки
            label = ttk.Label(
                window_setting_poduslugi, text=label_text, foreground="#d77991", font=("Trebuchet MS", 12, "bold"),
                background="#EEE2DC", anchor="center"
            )
            label.grid(row=2 + i, column=1)

            # Кнопки "перейти"
            btn = customtkinter.CTkButton(
                window_setting_poduslugi,
                text="Перейти",
                font=("Trebuchet MS", 14, "bold"),
                fg_color="#EDC7B7",
                hover_color="#A58B80",
                text_color="#AC3B61",
                command=lambda service_key=service_key: Thread(
                    target=self.start_window_poduslugi_name_btn, args=(service_key,)
                ).start()
            )
            btn.grid(row=2 + i, column=3)

    def save_path_name_btn_poduslugi(self,index_row, service_key, entry_var_full,entry_var_short):
        """Сохраняет данные в JSON файл."""
        file_path_json_path_btn = os.path.join(self.dir_path, self.sub_folder, f"ChooseUslugi{service_key}.json")
        with open(file_path_json_path_btn, "r", encoding="utf-8") as file:
            data_path_poduslugi_btn = json.load(file)
        data_path_poduslugi_btn[str(index_row)] = [entry_var_full,entry_var_short]
        with open(file_path_json_path_btn, "w", encoding="utf-8") as file:
            json.dump(data_path_poduslugi_btn, file, indent=4, ensure_ascii=False)


    def start_window_poduslugi_name_btn(self, service_key):     #Провaливает на уровень ниже настройки подуслуг
        
        file_path_json_path_btn = os.path.join(self.dir_path, self.sub_folder, f"ChooseUslugi{service_key}.json")
        with open(file_path_json_path_btn, "r", encoding='utf-8') as file:
            data_path_poduslugi_btn = json.load(file)

        window_setting_poduslugi = Toplevel(self.window_setting)
        window_setting_poduslugi.title("Настройки")
        window_setting_poduslugi.geometry("900x600")
        window_setting_poduslugi["bg"] = "#EEE2DC"
        for c in range(6):
            window_setting_poduslugi.columnconfigure(index=c, weight=1)
        for r in range(14):
            window_setting_poduslugi.rowconfigure(index=r, weight=1)

        label = ttk.Label(
            window_setting_poduslugi, text="Полный путь", foreground="#d77991", font=("Trebuchet MS", 12, "bold"),
            background="#EEE2DC", anchor="center"
        )
        label.grid(row=1, column=1)

        label = ttk.Label(
            window_setting_poduslugi, text="Отображение в таблицах итд", foreground="#d77991", font=("Trebuchet MS", 12, "bold"),
            background="#EEE2DC", anchor="center"
        )
        label.grid(row=1, column=2)

        for i, (key, value)  in enumerate(data_path_poduslugi_btn.items()):
            
            full_path = value[0]
            shortcut = value[1]
            # Поля ввода
            entry_var_full = customtkinter.StringVar()
            entry_var_full.set(full_path)  # Получаем начальное значение
        
            entry_full = customtkinter.CTkEntry(window_setting_poduslugi, textvariable=entry_var_full, width=300)
            entry_full.grid(row=2 + i, column=1)

            entry_var_short = customtkinter.StringVar()
            entry_var_short.set(shortcut)  # Получаем начальное значение
        
            entry_short = customtkinter.CTkEntry(window_setting_poduslugi, textvariable=entry_var_short, width=300)
            entry_short.grid(row=2 + i, column=2)

            # Кнопки "Сохранить"
            btn = customtkinter.CTkButton(
                window_setting_poduslugi,
                text="Сохранить",
                font=("Trebuchet MS", 14, "bold"),
                fg_color="#EDC7B7",
                hover_color="#A58B80",
                text_color="#AC3B61",
                command=lambda index_row=i+1, service_key=service_key, var_full=entry_var_full, var_short=entry_var_short: Thread(
                    target=self.save_path_name_btn_poduslugi, args=(index_row, service_key, var_full.get(), var_short.get())
                ).start()
            )
            btn.grid(row=2 + i, column=4)



    def ClockChooseUslugi(self):
        windowChooseUslugi = Toplevel(self.root)  
        windowChooseUslugi.title("Выбор целей")
        windowChooseUslugi.geometry("1600x500")
        windowChooseUslugi['bg'] = "#EEE2DC"
        for c in range(11): windowChooseUslugi.columnconfigure(index=c, weight=1)
        for r in range(10): windowChooseUslugi.rowconfigure(index=r, weight=1)
        
        get_selected_option = self.get_selected_option
        # Btn
        ConfirmBtnCU = customtkinter.CTkButton(windowChooseUslugi, text="Подтвердить",font=('Trebuchet MS', 16,'bold'),
                                               fg_color='#EDC7B7',hover_color='#A58B80',text_color='#AC3B61', 
                                               command=get_selected_option)
        ConfirmBtnCU.grid(row=9, column=6, ipadx=20, ipady=6)
        
        btnBack = customtkinter.CTkButton(windowChooseUslugi, text="Закрыть",font=('Trebuchet MS', 16,'bold'),
                                          fg_color='#EDC7B7',hover_color='#A58B80',text_color='#AC3B61', 
                                          command=lambda: windowChooseUslugi.destroy())
        btnBack.grid(row=9, column=4, ipadx=20, ipady=6)
        # MY
        label = ttk.Label(windowChooseUslugi, text="МУ", foreground="#d77991",style='ChoisUs.TLabel')
        label.grid(row=1, column=0)
        for i,UslugaMY in enumerate(self.uslugiMY):
            RadioMY = ttk.Radiobutton(windowChooseUslugi, text=UslugaMY["text"], style='Your.TRadiobutton', 
                                      value=UslugaMY["value"],variable=self.VarMY)
            RadioMY.grid(row=i + 2, column=0)

        # RVP
        label = ttk.Label(windowChooseUslugi, text="РВП", foreground="#d77991",style='ChoisUs.TLabel')
        label.grid(row=1, column=1)
        for i,UslugaRVP in enumerate(self.uslugiRVP):
            self.RadioRVP = ttk.Radiobutton(windowChooseUslugi, text=UslugaRVP["text"], 
                                            style='Your.TRadiobutton', value=UslugaRVP["value"],variable=self.VarRVP)
            self.RadioRVP.grid(row=i + 2, column=1)
            
        # VNJ
        label = ttk.Label(windowChooseUslugi, text="ВНЖ", foreground="#d77991",style='ChoisUs.TLabel')
        label.grid(row=1, column=2)
        
        for i,UslugaVNJ in enumerate(self.uslugiVNJ):
            RadioVNJ = ttk.Radiobutton(windowChooseUslugi, text=UslugaVNJ["text"], style='Your.TRadiobutton', value=UslugaVNJ["value"],variable=self.VarVNJ)
            RadioVNJ.grid(row=i + 2, column=2)
            
        # PATENT
        label = ttk.Label(windowChooseUslugi, text="Патент", foreground="#d77991",style='ChoisUs.TLabel')
        label.grid(row=1, column=3)
        
        for i,UslugaPATENT in enumerate(self.uslugiPATENT):
            RadioPATENT = ttk.Radiobutton(windowChooseUslugi, text=UslugaPATENT["text"], style='Your.TRadiobutton', value=UslugaPATENT["value"],variable=self.VarPATENT)
            RadioPATENT.grid(row=i + 2, column=3)
            
        # VIZA
        label = ttk.Label(windowChooseUslugi, text="Виза", foreground="#d77991",style='ChoisUs.TLabel')
        label.grid(row=1, column=4)
        
        for i,UslugaVIZA in enumerate(self.uslugiVIZA):
            RadioVIZA = ttk.Radiobutton(windowChooseUslugi, text=UslugaVIZA["text"], style='Your.TRadiobutton', value=UslugaVIZA["value"],variable=self.VarVIZA)
            RadioVIZA.grid(row=i + 2, column=4)
            
        # RP
        label = ttk.Label(windowChooseUslugi, text="РП", foreground="#d77991",style='ChoisUs.TLabel')
        label.grid(row=1, column=5)

        
        for i,UslugaRP in enumerate(self.uslugiRP):
            RadioRP = ttk.Radiobutton(windowChooseUslugi, text=UslugaRP["text"], style='Your.TRadiobutton', value=UslugaRP["value"],variable=self.VarRP)
            RadioRP.grid(row=i + 2, column=5)
            
        # ZPNEW
        label = ttk.Label(windowChooseUslugi, text="ЗП Нового", foreground="#d77991",style='ChoisUs.TLabel')
        label.grid(row=1, column=6)
        
        for i,UslugaZPNEW in enumerate(self.uslugiZPNEW):
            RadioZPNEW = ttk.Radiobutton(windowChooseUslugi, text=UslugaZPNEW["text"], style='Your.TRadiobutton', value=UslugaZPNEW["value"],variable=self.VarZPNEW)
            RadioZPNEW.grid(row=i + 2, column=6)
            
        # ZPOLD
        label = ttk.Label(windowChooseUslugi, text="ЗП Старого", foreground="#d77991",style='ChoisUs.TLabel')
        label.grid(row=1, column=7)
         
        for i, UslugaZPOLD in enumerate(self.uslugiZPOLD):
            RadioZPOLD = ttk.Radiobutton(windowChooseUslugi, text=UslugaZPOLD["text"], style='Your.TRadiobutton', value=UslugaZPOLD["value"],variable=self.VarZPOLD)
            RadioZPOLD.grid(row=i + 2, column=7)
            
        # ASI
        label = ttk.Label(windowChooseUslugi, text="АСИ", foreground="#d77991",style='ChoisUs.TLabel')
        label.grid(row=1, column=8)
         
        for i, UslugaASI in enumerate(self.uslugiASI):
            RadioASI = ttk.Radiobutton(windowChooseUslugi, text=UslugaASI["text"], style='Your.TRadiobutton', value=UslugaASI["value"],variable=self.VarASI)
            RadioASI.grid(row=i + 2, column=8)
            
        # RY
        label = ttk.Label(windowChooseUslugi, text="РУ", foreground="#d77991",style='ChoisUs.TLabel')
        label.grid(row=1, column=9)
        
        for i, UslugaRY in enumerate(self.uslugiRY):
            RadioRY = ttk.Radiobutton(windowChooseUslugi, text=UslugaRY["text"], style='Your.TRadiobutton', value=UslugaRY["value"],variable=self.VarRY)
            RadioRY.grid(row=i + 2, column=9)
            
        
            
    def ViewTable(self):
        self.windowViewTable = Toplevel(self.root)  
        self.windowViewTable.title("Таблица")
        self.windowViewTable.geometry("960x800")
        self.windowViewTable['bg'] = "#EEE2DC"
        style = ttk.Style()
        style.configure('Treeview.Heading', background="#F7F1EE",foreground='#123C69',font=('Trebuchet MS', 10))
        style.configure("Treeview", font=('Trebuchet MS', 11),background="#F7F1EE", fieldbackground="#F7F1EE",foreground='#123C69')
        
        self.columns = ("Наименование ОВД","МУ","РВП","ВНЖ","ПАТЕНТ","ВИЗА","РП","ЗП НОВЫЙ","ЗП СТАРЫЙ","ПРИГЛАШЕНИЕ","АСИ","РУ")
        self.tree = ttk.Treeview(self.windowViewTable,columns=self.columns, show='headings')
        self.tree.pack(padx=32,pady=32, fill=BOTH,expand=1)
        
        btnFrame = Frame(self.windowViewTable,width=120)
        btnFrame.pack(side=TOP)
        
        btnBack = customtkinter.CTkButton(btnFrame, text="Закрыть",font=('Trebuchet MS', 16,'bold'),fg_color='#EDC7B7',hover_color='#A58B80',text_color='#AC3B61', command=lambda: self.windowViewTable.destroy())
        btnBack.pack(side=LEFT)
        
        btnCreateExcel = customtkinter.CTkButton(btnFrame, text="Создать Excel",font=('Trebuchet MS', 16,'bold'),fg_color='#EDC7B7',hover_color='#A58B80',text_color='#AC3B61',command=lambda: self.CreateExcel())
        btnCreateExcel.pack(side=LEFT)
        
        btnFrame.place(relx=0.5, rely=0.98, anchor=CENTER)
        self.DataOVDs = [tuple(row) for row in self.df.itertuples(index=False)]
        self.DataOVDsYVM = [tuple(row) for row in self.dfMVD.itertuples(index=False)]
        print(self.DataOVDs)
        self.tree.heading(f"{self.columns[0]}",text=f"{self.columns[0]}")
        self.tree.column(f"#{0}",width=130)
        for i in range(1,len(self.columns)):         
            self.tree.heading(f"{self.columns[i]}",text=f"{self.columns[i]} ({self.NameSelectedUslugi[i-1]})")
            self.tree.column(f"#{i}",width=130)
            
            
            
        for i,DataOVD in enumerate(self.DataOVDs):
            self.tree.insert(parent='', index='end', values=(self.indexDF[i],)+DataOVD)
        self.tree.insert(parent='', index='end', values=('','','','','','','','','','','','',''))    
        for i,DataOVDYVM in enumerate(self.DataOVDsYVM):
            self.tree.insert(parent='', index='end', values=(self.indexDFMVD[i],)+DataOVDYVM)
        
        
    def CreateExcel(self):
        dataToDay = datetime.datetime.now()
        grey_fill = PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')
        blue_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        thins = Side(border_style="thin", color="000000")
        
        book = openpyxl.Workbook()
        book.remove(book.active)
        book.create_sheet('глубина очереди')
        
        ru_month_name = self.month_translation[calendar.month_name[dataToDay.month]]
        book.active.append((f"{dataToDay.day} {ru_month_name} {dataToDay.year}",'', '', '', '', '', '', '', '', '', '', ''))
        book.active.merge_cells('A1:L1')
        for cell in book.active[book.active.max_row]:
            cell.fill = grey_fill
        book.active.append(('  Вид услуги',f"МУ ({self.NameSelectedUslugi[0]})",f"РВП ({self.NameSelectedUslugi[1]})"
                                  ,f"ВНЖ ({self.NameSelectedUslugi[2]})",
                                  f"ПАТЕНТ ({self.NameSelectedUslugi[3]})",
                                 f"ВИЗА ({self.NameSelectedUslugi[4]})",
                                  f'РП ({self.NameSelectedUslugi[5]})',f"ЗП НОВОГО ОБРАЗЦА ({self.NameSelectedUslugi[7]})"
                                  ,f"ЗП СТАРОГО ОБРАЗЦА ({self.NameSelectedUslugi[7]})",f"ПРИГЛАШЕНИЕ ({self.NameSelectedUslugi[8]})",
                                    f"АСИ ({self.NameSelectedUslugi[9]})",
                                  f"РУ гр. РФ по месту жит-ва ({self.NameSelectedUslugi[10]})")) 
        book.active.append(('  Наименование ОВД',1,2,3,4,5,6,7,8,9,10,11))
        for cell in book.active[book.active.max_row]:
            cell.fill = blue_fill
        for i,row in enumerate(self.DataOVDs):
            book.active.append((f"  {self.indexDF[i]}",)+row)
            
        book.active.append(('', '', '', '', '', '', '', '', '', '', ''))
        
        for cell in book.active[book.active.max_row]:
            cell.fill = blue_fill
        first_cell = book.active.cell(row=book.active.max_row, column=1)
        last_cell = book.active.cell(row=book.active.max_row, column=12)
        cell_range = f'{first_cell.coordinate}:{last_cell.coordinate}'
        book.active.merge_cells(cell_range)
        for i,row in enumerate(self.DataOVDsYVM):
            book.active.append((f"  {self.indexDFMVD[i]}",)+row)
        for col_index in range(1, 13):
            column_letter = get_column_letter(col_index)
            book.active.column_dimensions[column_letter].width = 15
            book.active[f"{column_letter}{2}"].alignment = Alignment(wrap_text=True,horizontal='center',vertical="center") 
            book.active[f"{column_letter}{3}"].alignment = Alignment(horizontal='center') 
            book.active[f"{column_letter}{2}"].font= Font(color='0070C0')
        book.active.column_dimensions['A'].width = 28
        book.active.row_dimensions[2].height = 100
        
        for row in book.active.iter_rows():
            for cell in row:
                cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)
        
        for i, row in enumerate(book.active.iter_rows(values_only=True)):
            for j, cell in enumerate(row):
                column_letter = get_column_letter(j+1)
                book.active[f"{column_letter}{i+1}"].font = Font(size=14,bold=True)
                if i>2 and j>0:
                    book.active[f"{column_letter}{i+1}"].alignment = Alignment(horizontal='center',vertical="center")
                    if isinstance(cell, str) == True:
                        if '(' in cell:
                            book.active[f"{column_letter}{i+1}"].font = Font(color='FF0000',size=14,bold=True)
                            book.active[f"{column_letter}{i+1}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                            book.active[f"A{i+1}"].font = Font(color='FF0000',size=14,bold=True)
                if cell == 'X':
                    book.active[f"{column_letter}{i+1}"].font = Font(color='0070C0',size=14,bold=True)      
                if cell == 'Х':
                    book.active[f"{column_letter}{i+1}"].font = Font(color='FF0000',size=14,bold=True)
                    book.active[f"{column_letter}{i+1}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                    book.active[f"A{i+1}"].font = Font(color='FF0000',size=14,bold=True)
                if cell == '_':
                    book.active[f"{column_letter}{i+1}"].font = Font(color='FF0000',size=14,bold=True)
                    book.active[f"{column_letter}{i+1}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                    book.active[f"A{i+1}"].font = Font(color='FF0000',size=14,bold=True)
                    
        book.save(f"Таблица глубины очереди {dataToDay.day}.{dataToDay.month}.xlsx")
    def StartOfIndividualServices(self):
        windowIndividualServices = Toplevel(self.root)  
        windowIndividualServices["bg"] = "#EEE2DC"
        windowIndividualServices.title("Выбор отдельного запуска услуг")
        windowIndividualServices.geometry("650x449")
        
        
        
        for c in range(8): windowIndividualServices.columnconfigure(index=c, weight=1)
        for r in range(11): windowIndividualServices.rowconfigure(index=r, weight=1)
            
        
        btnBack = customtkinter.CTkButton(windowIndividualServices, text="Закрыть",font=('Trebuchet MS', 16,'bold'),
                                          fg_color='#EDC7B7',hover_color='#A58B80',text_color='#AC3B61', command=lambda: windowIndividualServices.destroy())
        btnBack.grid(row=10, column=3, ipadx=20, ipady=6,columnspan=2)
        
        label = ttk.Label(windowIndividualServices, text="Мигрант услуги", foreground="#d77991", style='Your.TLabel')
        label.grid(row=1, column=1)
        for i,UslugaMY in enumerate(self.nameovdlistru[:5]):
            RadioMY = ttk.Checkbutton(windowIndividualServices, style='Your.TCheckbutton',text=self.nameovdlistru[i],
                                      variable=self.listRadioStringVar[i],onvalue=f"{UslugaMY}")
            RadioMY.grid(row=i + 2, column=1)                                          
        btnStartBrowser = customtkinter.CTkButton(windowIndividualServices, text="Запустить услуги",
                                                  font=('Trebuchet MS', 16,'bold'),fg_color='#EDC7B7',hover_color='#A58B80',text_color='#AC3B61',command=lambda: Thread(target=self.GetStartIndividualServicesImigrant).start())
        btnStartBrowser.grid(row=9, column=1, ipadx=20, ipady=6)
        
        label = ttk.Label(windowIndividualServices, text="Гражданин услуги", foreground="#d77991", style='Your.TLabel')
        label.grid(row=1, column=6)
        for i,UslugaMY in enumerate(self.nameovdlistru[5:]):
            RadioMY = ttk.Checkbutton(windowIndividualServices, text=self.nameovdlistru[i+5], 
                                      style='Your.TCheckbutton',variable=self.listRadioStringVar[i+5],onvalue=f"{UslugaMY}")
            RadioMY.grid(row=i + 2, column=6)
        btnStartBrowser = customtkinter.CTkButton(windowIndividualServices, text="Запустить услуги",
                                                  font=('Trebuchet MS', 16,'bold'),fg_color='#EDC7B7',
                                                  hover_color='#A58B80',text_color='#AC3B61',command=lambda: Thread(target=self.GetStartIndividualServicesRP).start())
        btnStartBrowser.grid(row=9, column=6, ipadx=20, ipady=6)
    
    def GetStartIndividualServicesRP(self):
        for i, var in enumerate(self.listRadioStringVar[5:]):
            if var.get() == self.nameovdlistru[i+5]:
                self.AllProgram[i+5]()
    
    def GetStartIndividualServicesImigrant(self):
        for i, var in enumerate(self.listRadioStringVar[:5]):
            if var.get() == self.nameovdlistru[i]:
                self.AllProgram[i]()

  
    def get_selected_option(self):
        # Создаем временные списки для новых значений
        temp_names = []
        temp_values = []
    
        # Собираем выбранные значения для всех 11 динамических категорий
        # (исключая "Приглашение" под индексом 8)
        for i, get in enumerate(self.listRadioVar):
            # Пропускаем индекс 8 (Приглашение) - он статичный
            if i == 8:
                continue
        
            for dic in self.AllUslugi[i]:
                if dic['value'] == get.get():
                    temp_names.append(dic['text'])
                    temp_values.append(dic['value'])
                    break  # Нашли нужный элемент, выходим из внутреннего цикла
    
        # Вставляем статичное значение "Приглашение" на позицию 8
        # Нужно получить его из self.uslugiINVITATION
        invitation_text = ""
        invitation_value = ""
        for dic in self.uslugiINVITATION:
            if dic['value'] == self.VarINVITATION.get():
                invitation_text = dic['text']
                invitation_value = dic['value']
                break
    
        # Вставляем на позицию 8 (индекс 8)
        temp_names.insert(8, invitation_text)
        temp_values.insert(8, invitation_value)
    
        # Проверяем, что все 11 категорий были заполнены
        # Должно быть: 10 динамических + 1 статическое = 11
        while len(temp_names) < 11:
            temp_names.append("Не выбрано")
            temp_values.append("")
    
        # Обновляем основные списки
        self.NameSelectedUslugi = temp_names
        self.FullValueSelectedUslugi = temp_values
    
        # Сохраняем в JSON файл
        sub_folder = "files"
        file_path = os.path.join(self.dir_path, sub_folder, "SelectedUslugi.json")
    
        # Создаем список словарей для сохранения
        selected_data = []
        for i in range(len(self.NameSelectedUslugi)):
            selected_data.append({
                "name": self.NameSelectedUslugi[i],
                "value": self.FullValueSelectedUslugi[i]
            })
    
        # Записываем в JSON файл
        with open(file_path, "w", encoding='utf-8') as file:
            json.dump(selected_data, file, ensure_ascii=False, indent=4)
    
        self.create_data_frames()    
    def rightEmailTelepthoneRegoin(self):
        emailRight = WebDriverWait(self.browser, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME,"button")))
        emailRight.click()
        time.sleep(4)

        telephone = (By.XPATH,"//button[@type='button']")
        telefoneRight = WebDriverWait(self.browser, 20).until(
            EC.visibility_of_element_located(telephone))
        WebDriverWait(self.browser, 20).until(EC.element_to_be_clickable(telephone))
        telefoneRight.click()

        regionSearch = WebDriverWait(self.browser, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME,"search-input")))
        regionSearch.clear()
        regionSearch.send_keys("Новосибирская обл")

        region = WebDriverWait(self.browser, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME,"highlighted")))
        region.click()

        right = WebDriverWait(self.browser, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME,"button")))
        right.click()
    def StartReplasementPassport(self):
        StartButton = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME,"button")))

        StartButton.click()

        buttonReplacementPassport = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.data_path_btn['rp']}')]")))
        buttonReplacementPassport.click()

        right = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.FullValueSelectedUslugi[5]}')]")))
        right.click()
        self.rightEmailTelepthoneRegoin()         
    def StartForeignPassportOfANewType(self):
        StartButton = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME,"button")))

        StartButton.click()

        ForeignPassport = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.data_path_btn['zp']}')]")))
        ForeignPassport.click()

        NewType = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,"//span[contains(text(),'Нового образца')]")))
        NewType.click()

        Me = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.FullValueSelectedUslugi[6]}')]")))
        Me.click()

        self.rightEmailTelepthoneRegoin()
    def StartForeignPassportOfAOldType(self):
        StartButton = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME,"button")))

        StartButton.click()

        ForeignPassport = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.data_path_btn['zp']}')]")))
        ForeignPassport.click()

        NewType = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,"//span[contains(text(),'Старого образца')]")))
        NewType.click()

        Me = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.FullValueSelectedUslugi[7]}')]")))
        Me.click()

        self.rightEmailTelepthoneRegoin()
    def StartInvitation(self):
        StartButton = WebDriverWait(self.browser, 20).until(
           EC.presence_of_element_located((By.CLASS_NAME,"button")))

        StartButton.click()

        ForeignPassport = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.data_path_btn['invitation']}')]")))
        ForeignPassport.click()

        self.rightEmailTelepthoneRegoin()
    def StartASI(self):
        StartButton = WebDriverWait(self.browser, 20).until(
                EC.presence_of_element_located((By.CLASS_NAME,"button")))

        StartButton.click()

        ASI = WebDriverWait(self.browser, 20).until(
                EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.data_path_btn['asi']}')]")))
        ASI.click()

        ASIAboutAnotherPerson = WebDriverWait(self.browser, 20).until(
                EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.FullValueSelectedUslugi[9]}')]")))
        ASIAboutAnotherPerson.click()

        self.rightEmailTelepthoneRegoin() 
            
    def StartRegistrationAtPlaceOfResidence(self):
        StartButton = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME,"button")))

        StartButton.click()

        RegistrationAtPlaceOfResidence = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.data_path_btn['ry']}')]")))
        RegistrationAtPlaceOfResidence.click()

        RegistrationAtPlaceOfResidenceRegister = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.FullValueSelectedUslugi[10]}')]")))
        RegistrationAtPlaceOfResidenceRegister.click()

        self.rightEmailTelepthoneRegoin()
    #def StartGPPS(self):
    #    StartButton = WebDriverWait(self.browser, 20).until(
    #        EC.presence_of_element_located((By.CLASS_NAME,"button")))
    #
    #    StartButton.click()
    #
    #    GPPS = WebDriverWait(self.browser, 20).until(
    #        EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.data_path_btn['gpps']}')]")))
    #    GPPS.click()
    #
    #    GPPSGet = WebDriverWait(self.browser, 20).until(
    #        EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.FullValueSelectedUslugi[11]}')]")))
    #    GPPSGet.click()
    #
    #    self.rightEmailTelepthoneRegoin()    
    def StartMY(self):
        StartButton = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME,"button")))

        StartButton.click()

        MY = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.data_path_btn['my']}')]")))
        MY.click()

        MYGet = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.FullValueSelectedUslugi[0]}')]")))
        MYGet.click()
    
        self.rightEmailTelepthoneRegoin() 
    def StartVNJ(self):
        StartButton = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME,"button")))

        StartButton.click()

        VNJ = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.data_path_btn['vnj']}')]")))
        VNJ.click()

        VNJGet = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.FullValueSelectedUslugi[2]}')]")))
        VNJGet.click()

        self.rightEmailTelepthoneRegoin()
    def StartRVP(self):
        StartButton = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME,"button")))

        StartButton.click()

        RVP = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.data_path_btn['rvp']}')]")))
        RVP.click()

        RVPGet = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.FullValueSelectedUslugi[1]}')]")))
        RVPGet.click()

        self.rightEmailTelepthoneRegoin()
    def StartPATENT(self):
        StartButton = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME,"button")))

        StartButton.click()

        PATENT = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.data_path_btn['patent']}')]")))
        PATENT.click()

        PATENTGet = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.FullValueSelectedUslugi[3]}')]")))
        PATENTGet.click()

        self.rightEmailTelepthoneRegoin()
    def StartVIZA(self):
        StartButton = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME,"button")))

        StartButton.click()

        VIZA = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.data_path_btn['viza']}')]")))
        VIZA.click()

        VIZAGet = WebDriverWait(self.browser, 20).until(
            EC.presence_of_element_located((By.XPATH,f"//span[contains(text(),'{self.FullValueSelectedUslugi[4]}')]")))
        VIZAGet.click()

        self.rightEmailTelepthoneRegoin()
        
    def WorkDaysAfterToDay(self,today,YVM):
        data_today = datetime.datetime.now().date()
        copy_today = today
        WorkDays = 0
        DaysInMonth = calendar.monthrange(data_today.year, data_today.month)[1]
        day_violation = int(today)
        data_today += datetime.timedelta(days=1)
        
        if today < data_today.day:
            today+=DaysInMonth
        
        while today >= data_today.day:
            DaysInMonth = calendar.monthrange(data_today.year, data_today.month)[1]
            if data_today.day == DaysInMonth:
                today-=DaysInMonth
                
            if YVM==False:
                if data_today not in self.holidays:
                    if data_today.weekday() == 0 or data_today.weekday() == 1 or data_today.weekday() == 3 or data_today.weekday() == 4 or data_today.weekday() == 5:  
                        WorkDays += 1
            if YVM==True:
                if data_today not in self.holidays:
                    if data_today.weekday() == 0 or data_today.weekday() == 1 or data_today.weekday() == 3 or data_today.weekday() == 4:  
                        WorkDays += 1
            data_today += datetime.timedelta(days=1)
            
            if WorkDays > 5:
                day_violation = f'{copy_today}({WorkDays-5})'
        return day_violation    
    def SearchOVDSpeedUp(self,ListOVD,NameService, StartUsluga, NameOVD,indexDF,df,YVM=False):
        for i in range(len(NameOVD)):
            if NameOVD[i] not in ListOVD:
                new_data = "X"
                index_to_insert = indexDF[i]

                df.at[index_to_insert, NameService] = new_data

                continue
            try:
                
                search_input = (By.CLASS_NAME,"search-input")
                self.wait.until(EC.visibility_of_element_located(search_input))
                rightRegionSearch = self.wait.until(EC.element_to_be_clickable(search_input))

                rightRegionSearch.click()
                rightRegionSearch.clear()
                rightRegionSearch.send_keys(NameOVD[i])

                time.sleep(1.5)

            except ElementClickInterceptedException:
                try:    
                    time.sleep(3.5)
                    rightRegionSearch = self.wait.until(EC.visibility_of_element_located(search_input))

                    rightRegionSearch.click()
                    rightRegionSearch.clear()
                    rightRegionSearch.send_keys(NameOVD[i])
                except ElementClickInterceptedException:
                    try:
                        time.sleep(3.5)
                        rightRegionSearch = self.wait.until(EC.visibility_of_element_located(search_input))

                        rightRegionSearch.click()
                        rightRegionSearch.clear()
                        rightRegionSearch.send_keys(NameOVD[i])
                    except ElementClickInterceptedException:
                        self.BackMainPage()
                        StartUsluga()
                        continue
            except TimeoutException:
                print("Ошибка страница не загрузилась")
                break
                
            try :
                time.sleep(1)
                butLinkOnSearchMap = (By.CLASS_NAME,"balloon-btn") 
                ButtonLinkOnSearchMap = self.wait.until(EC.visibility_of_element_located(butLinkOnSearchMap))
                ButtonLinkOnSearchMap.click() # нажатие на название подразделение после чего появляется кнопка


            except TimeoutException:
                try: 
                    butNotFound = (By.CLASS_NAME,"not-found-text")
                    self.wait.until(EC.visibility_of_element_located(butNotFound))
                    new_data = "Х"
                    index_to_insert = indexDF[i]
                    df.at[index_to_insert, NameService] = new_data
                    time.sleep(1)
                    continue
                except:
                    print("Неизвестная ошибка")
                    rightRegionSearch.clear()
                    continue

            chooseBut = By.CLASS_NAME,"wide"    
            self.wait.until(EC.visibility_of_element_located(chooseBut))
            chooseButton = self.wait.until(EC.element_to_be_clickable(chooseBut)) # нажатие на кнопку
            time.sleep(1)
            chooseButton.click()

            kalendNum = (By.XPATH,"//epgu-cf-ui-constructor-screen-pad[@class='ng-star-inserted']") 
            try:
                self.wait.until(EC.visibility_of_all_elements_located(kalendNum))
                kalendar = self.browser.find_elements(By.CLASS_NAME,"locked")
                soup = BeautifulSoup(self.browser.page_source, "lxml")
                dataDay = []

                for day in kalendar:
                    dayNumber = day.find_element(By.CLASS_NAME,"calendar-day-text").text
                    dataDay.append(dayNumber)


                dataDaySet = frozenset(dataDay)
                dayRecord = [item for item in  self.AllDayinMonth if item not in  dataDaySet] 
                dayRecord = dayRecord[0]
                new_data = int(dayRecord)
                
                dayRecordVer = self.WorkDaysAfterToDay(new_data,YVM)
                
                index_to_insert = indexDF[i]
                df.at[index_to_insert, NameService] = str(dayRecordVer)

                print(dayRecord)
            except TimeoutException:
                TimeOutKalendar = (By.CLASS_NAME,"button") 
                try:
                    self.wait.until(EC.visibility_of_all_elements_located(TimeOutKalendar))
                    self.browser.refresh()
                    print("_")
                    new_data = "_"
                    index_to_insert = indexDF[i]
                    df.at[index_to_insert, NameService] = new_data
                    StartUsluga()
                    continue
                except TimeoutException:
                    self.BackMainPage()
                    StartUsluga()
                    continue
            try:
                backFromCalendar = self.browser.find_element(By.CLASS_NAME,"link-btn")
                backFromCalendar.click()
                time.sleep(1)

            except ElementClickInterceptedException:
                try:
                    backIfNotHaveRecord = self.browser.find_element(By.CSS_SELECTOR,"lib-button.conf-modal__button:nth-child(2) > div:nth-child(1) > button:nth-child(1)")
                    new_data = "_"
                    index_to_insert = indexDF[i]
                    df.at[index_to_insert, NameService] = new_data
                    backIfNotHaveRecord.click()
                    time.sleep(2)
                except NoSuchElementException :
                    #refreshPage = self.browser.find_element(By.XPATH,"//span[contains(text(),'Попробовать ещё раз')]") 
                    new_data = "-"
                    index_to_insert = indexDF[i]
                    df.at[index_to_insert, NameService] = new_data
                    StartUsluga()
                    continue
        return df
    def BackMainPage(self):
        self.browser.get('https://www.gosuslugi.ru/600300/1/form')
        time.sleep(1)
        
    def StartUslugaInProgram(self,StartUsluga):
        self.BackMainPage()
        try:
            StartUsluga()
            
        except TimeoutException:
            self.BackMainPage()
            try:
                StartUsluga()
            except TimeoutException:
                return "Ошибка загрузки страницы"
    def replasementPasport(self):
        self.StartUslugaInProgram(self.StartReplasementPassport)
        
        self.df = self.SearchOVDSpeedUp(ListOVD=self.NameOVDReplasement, NameService=f"РП ({self.NameSelectedUslugi[5]})",
                              StartUsluga=self.StartReplasementPassport, NameOVD=self.NameOVD, indexDF=self.indexDF,
                              df=self.df)
        self.dfMVD = self.SearchOVDSpeedUp(ListOVD=self.NameOVDReplasementMVD, NameService=f"РП ({self.NameSelectedUslugi[5]})",
                              StartUsluga=self.StartReplasementPassport, NameOVD=self.NameOVDMVD,
                              indexDF=self.indexDFMVD, df=self.dfMVD,YVM=True)
    def ForeignPassportOfANewType(self):
        self.StartUslugaInProgram(self.StartForeignPassportOfANewType)
        
        self.df = self.SearchOVDSpeedUp(ListOVD=self.NameOVDPassNewType, NameService=f"ЗП НОВОГО ОБРАЗЦА ({self.NameSelectedUslugi[6]})",
                              StartUsluga=self.StartForeignPassportOfANewType, NameOVD=self.NameOVD, indexDF=self.indexDF,
                              df=self.df)
        self.dfMVD = self.SearchOVDSpeedUp(ListOVD=self.NameOVDPassNewTypeMVD, NameService=f"ЗП НОВОГО ОБРАЗЦА ({self.NameSelectedUslugi[6]})",
                              StartUsluga=self.StartForeignPassportOfANewType, NameOVD=self.NameOVDMVD,
                              indexDF=self.indexDFMVD, df=self.dfMVD,YVM=True)
    def ForeignPassportOfAOldType(self):
        self.StartUslugaInProgram(self.StartForeignPassportOfAOldType)
        
        self.df = self.SearchOVDSpeedUp(ListOVD=self.NameOVDPassOldType, NameService=f"ЗП СТАРОГО ОБРАЗЦА ({self.NameSelectedUslugi[7]})",
                                  StartUsluga=self.StartForeignPassportOfAOldType, NameOVD=self.NameOVD, indexDF=self.indexDF,
                                  df=self.df)
        self.dfMVD = self.SearchOVDSpeedUp(ListOVD=self.NameOVDPassOldTypeMVD, NameService=f"ЗП СТАРОГО ОБРАЗЦА ({self.NameSelectedUslugi[7]})",
                                  StartUsluga=self.StartForeignPassportOfAOldType, NameOVD=self.NameOVDMVD, indexDF=self.indexDFMVD,
                                  df=self.dfMVD,YVM=True)
    def Invitation(self):
        self.StartUslugaInProgram(self.StartInvitation)
        
        self.df = self.SearchOVDSpeedUp(ListOVD=self.NameOVDInvitation, NameService=f"ПРИГЛАШЕНИЕ ({self.NameSelectedUslugi[8]})", StartUsluga=self.StartInvitation, NameOVD=self.NameOVD, indexDF=self.indexDF,
                              df=self.df)
        self.dfMVD = self.SearchOVDSpeedUp(ListOVD=self.NameOVDInvitationMVD, NameService=f"ПРИГЛАШЕНИЕ ({self.NameSelectedUslugi[8]})", StartUsluga=self.StartInvitation,
                              NameOVD=self.NameOVDMVD, indexDF=self.indexDFMVD, df=self.dfMVD,YVM=True)
    def ASI(self):
        self.StartUslugaInProgram(self.StartASI)
        
        self.df = self.SearchOVDSpeedUp(ListOVD=self.NameOVDASI, NameService=f"АСИ ({self.NameSelectedUslugi[9]})", StartUsluga=self.StartASI, NameOVD=self.NameOVD, indexDF=self.indexDF,
                              df=self.df)
        self.dfMVD = self.SearchOVDSpeedUp(ListOVD=self.NameOVDASIMVD, NameService=f"АСИ ({self.NameSelectedUslugi[9]})", StartUsluga=self.StartASI, NameOVD=self.NameOVDMVD,
                              indexDF=self.indexDFMVD, df=self.dfMVD,YVM=True)
    def RegistrationAtPlaceOfResidence(self):
        self.StartUslugaInProgram(self.StartRegistrationAtPlaceOfResidence)
        
        self.df = self.SearchOVDSpeedUp(ListOVD=self.NameOVDRegistrationAtPlaceOfResidence, NameService=f"РУ гр. РФ по месту жит-ва ({self.NameSelectedUslugi[10]})",
                              StartUsluga=self.StartRegistrationAtPlaceOfResidence, NameOVD=self.NameOVD, indexDF=self.indexDF,
                              df=self.df)
        self.dfMVD = self.SearchOVDSpeedUp(ListOVD=self.NameOVDRegistrationAtPlaceOfResidenceMVD, NameService=f"РУ гр. РФ по месту жит-ва ({self.NameSelectedUslugi[10]})",
                              StartUsluga=self.StartRegistrationAtPlaceOfResidence, NameOVD=self.NameOVDMVD, indexDF=self.indexDFMVD,
                              df=self.dfMVD,YVM=True)
    #def GPPS(self):
    #    self.StartUslugaInProgram(self.StartGPPS)
    #    
    #    self.df = self.SearchOVDSpeedUp(ListOVD=self.NameOVDGPPS, NameService=f"ГППС ({self.NameSelectedUslugi[11]})", StartUsluga=self.StartGPPS, NameOVD=self.NameOVD, indexDF=self.indexDF,
    #                          df=self.df)
    #    self.dfMVD = self.SearchOVDSpeedUp(ListOVD=self.NameOVDGPPSMVD, NameService=f"ГППС ({self.NameSelectedUslugi[11]})", StartUsluga=self.StartGPPS, NameOVD=self.NameOVDMVD,
    #                          indexDF=self.indexDFMVD, df=self.dfMVD,YVM=True)
    def MY(self):
        self.StartUslugaInProgram(self.StartMY)
        
        self.df = self.SearchOVDSpeedUp(ListOVD=self.NameOVDMY, NameService=f"МУ ({self.NameSelectedUslugi[0]})", StartUsluga=self.StartMY, NameOVD=self.NameOVD, indexDF=self.indexDF,
                              df=self.df)
        self.dfMVD = self.SearchOVDSpeedUp(ListOVD=self.NameOVDMYMVD, NameService=f"МУ ({self.NameSelectedUslugi[0]})", StartUsluga=self.StartMY, NameOVD=self.NameOVDMVD,
                              indexDF=self.indexDFMVD, df=self.dfMVD,YVM=True)
    def RVP(self):
        self.StartUslugaInProgram(self.StartRVP)
        
        self.df = self.SearchOVDSpeedUp(ListOVD=self.NameOVDRVP, NameService=f"РВП ({self.NameSelectedUslugi[1]})", StartUsluga=self.StartRVP, NameOVD=self.NameOVD, indexDF=self.indexDF,
                              df=self.df)
        self.dfMVD = self.SearchOVDSpeedUp(ListOVD=self.NameOVDRVPMVD, NameService=f"РВП ({self.NameSelectedUslugi[1]})", StartUsluga=self.StartRVP, NameOVD=self.NameOVDMVD,
                              indexDF=self.indexDFMVD, df=self.dfMVD,YVM=True)
    def VNJ(self):
        self.StartUslugaInProgram(self.StartVNJ)
        
        self.df = self.SearchOVDSpeedUp(ListOVD=self.NameOVDVNJ, NameService=f"ВНЖ ({self.NameSelectedUslugi[2]})", StartUsluga=self.StartVNJ, NameOVD=self.NameOVD, indexDF=self.indexDF,
                              df=self.df)
        self.dfMVD = self.SearchOVDSpeedUp(ListOVD=self.NameOVDVNJMVD, NameService=f"ВНЖ ({self.NameSelectedUslugi[2]})", StartUsluga=self.StartVNJ, NameOVD=self.NameOVDMVD,
                              indexDF=self.indexDFMVD, df=self.dfMVD,YVM=True)
    def PATENT(self):
        self.StartUslugaInProgram(self.StartPATENT)
        
        self.df = self.SearchOVDSpeedUp(ListOVD=self.NameOVDPATENT, NameService=f"ПАТЕНТ ({self.NameSelectedUslugi[3]})", StartUsluga=self.StartPATENT, NameOVD=self.NameOVD, indexDF=self.indexDF,
                              df=self.df)
        self.dfMVD = self.SearchOVDSpeedUp(ListOVD=self.NameOVDPATENTMVD, NameService=f"ПАТЕНТ ({self.NameSelectedUslugi[3]})", StartUsluga=self.StartPATENT, NameOVD=self.NameOVDMVD,
                              indexDF=self.indexDFMVD, df=self.dfMVD,YVM=True)  
    def VIZA(self):
        self.StartUslugaInProgram(self.StartVIZA)
        
        self.df = self.SearchOVDSpeedUp(ListOVD=self.NameOVDVIZA, NameService=f"ВИЗА ({self.NameSelectedUslugi[4]})", StartUsluga=self.StartVIZA, NameOVD=self.NameOVD, indexDF=self.indexDF,
                              df=self.df)
        self.dfMVD = self.SearchOVDSpeedUp(ListOVD=self.NameOVDVIZAMVD, NameService=f"ВИЗА ({self.NameSelectedUslugi[4]})", StartUsluga=self.StartVIZA, NameOVD=self.NameOVDMVD,
                              indexDF=self.indexDFMVD, df=self.dfMVD,YVM=True)      
    def StartProgramRP(self):
        self.replasementPasport()
        self.ForeignPassportOfANewType()
        self.ForeignPassportOfAOldType()
        self.Invitation()
        self.ASI()
        self.RegistrationAtPlaceOfResidence()

    def StartProgramImigrant(self):
        self.MY()
        self.RVP()
        self.VNJ()
        self.PATENT()
        self.VIZA()
if __name__ == "__main__":
    window = Window()
    window.run()